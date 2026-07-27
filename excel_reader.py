# -*- coding: utf-8 -*-
"""고객 엑셀 파싱 - C열(인스타그램 URL) / F열(개인화 DM 문구)만 읽는다.

고객 시트는 A열부터 이어지지 않고(연속 아님) 컬럼이 더 있다("섭외메시지 캡처" 등).
그 컬럼들은 매크로 대상이 아니므로 절대 건드리지 않고 C/F 두 칸만 읽는다.

헤더 행 유무를 사용자에게 물어보지 않고 자동으로 걸러낸다: C열 값이 실제 인스타그램
URL 처럼 생긴 행만 데이터로 취급한다(그 외 - 헤더 라벨, 빈 줄, "URL" 같은 제목 - 는
조용히 건너뛴다). F열은 있는 그대로(문구를 임의로 다듬거나 템플릿화하지 않는다).
"""

import re
from dataclasses import dataclass

from openpyxl import load_workbook

_URL_RE = re.compile(r"instagram\.com/[A-Za-z0-9._][A-Za-z0-9._/]*", re.IGNORECASE)

COL_URL = 3   # C열 (1-based)
COL_MSG = 6   # F열 (1-based)


@dataclass
class Row:
    row_no: int      # 실제 엑셀 행 번호(1-based) - 진행상황 키로 사용
    url: str
    username: str
    message: str


def _extract_username(url):
    """인스타그램 프로필 URL에서 사용자명만 뽑는다. instagram.com/<username>/..."""
    m = _URL_RE.search(url)
    if not m:
        return None
    tail = m.group(0).split("instagram.com/", 1)[-1]
    username = tail.strip("/").split("/")[0].split("?")[0]
    if not username or username.lower() in ("p", "reel", "reels", "explore", "stories", "direct"):
        return None
    return username


def load_rows(path):
    """엑셀에서 (C=URL, F=메시지) 데이터 행만 뽑아 Row 리스트로 반환.

    - C열이 인스타그램 프로필 URL 처럼 생기지 않은 행(헤더/빈줄/무관한 값)은 건너뜀.
    - F열이 비어있으면 그 행은 스킵 대상으로 표시(팔로우만 하고 DM 생략은 하지 않는다;
      스펙상 "F열 DM 내용" 이 필수이므로 없으면 해당 사람 전체를 스킵하고 경고만 남긴다).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    skipped_no_message = []
    for row_no, row in enumerate(ws.iter_rows(min_col=1, max_col=max(COL_URL, COL_MSG)), start=1):
        raw_url = row[COL_URL - 1].value if len(row) >= COL_URL else None
        raw_msg = row[COL_MSG - 1].value if len(row) >= COL_MSG else None
        url = str(raw_url).strip() if raw_url is not None else ""
        message = str(raw_msg).strip() if raw_msg is not None else ""
        if not url:
            continue
        username = _extract_username(url)
        if not username:
            continue  # 헤더/설명 텍스트/빈 값 등 - 실제 프로필 URL 이 아님
        if not message:
            skipped_no_message.append(row_no)
            continue
        rows.append(Row(row_no=row_no, url=url, username=username, message=message))
    wb.close()
    return rows, skipped_no_message
