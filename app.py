# -*- coding: utf-8 -*-
"""인스타 DM 매크로 - Tkinter GUI.

화면 구성:
  0) 버전/업데이트 줄: 지금 버전 + [지금 업데이트] + [다운로드 주소 복사]. 자동 교체가
     막히면 이 줄이 노란 경고로 바뀌고 사유와 수동 다운로드 주소를 그대로 보여 준다.
     (v1.7.0. 고객 1898680 이 두 번 - 1.3.1, 1.5.0 - 옛 버전에 '조용히' 묶였다.
      막힌 사실이 화면에 보이지 않는 상태를 없애는 것이 이 줄의 목적이다.)
  1) 계정: 별명 드롭다운(저장된 별명 + 그 별명에 묶인 인스타 아이디) + [+ 새 별명]
     + [계정 기록 지우기] + [로그인/계정 전환] + [로그아웃]. 로그인은 항상 실제로 뜬 크롬
     창에서 사람이 직접 아이디/비번을 입력한다(프로그램은 로그인 여부만 감지). 별명마다
     크롬 프로필이 갈린다. 크롬 창에서 인스타 자체 '계정 전환'을 써도 되고, 그때 실제로
     어느 계정으로 도는지는 '현재 실행 계정' 줄에 항상 표시된다.

     v1.6.0 규칙 두 가지(고객 1898680 이 다섯 번 리포트한 그 문제):
       - **고객이 로그인한 계정에서 프로그램이 절대 끌어내지 않는다.** 별명에 저장된 계정과
         크롬 창의 계정이 다르면 저장값을 고쳐 쓴다(브라우저를 고치지 않는다). 인스타 자체
         계정 전환을 프로그램이 대신 눌러 주는 기능은 체크박스로 **꺼진 채** 제공된다.
       - **[시작] 은 플래그가 아니라 살아 있는 크롬 창을 보고 판단한다.** 로그인 확인이 아직
         돌고 있으면 '로그인하세요' 팝업 대신 기다렸다가 자동으로 시작한다.
  2) 엑셀 파일: [찾아보기] - C열(URL)/F열(DM 문구)만 읽는다.
  3) [시작] / [중지] / [진행상황 초기화]
  4) 로그 창(스크롤) + 진행 통계 라벨
"""

import os
import random
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import bridge
import config
import excel_reader
import progress_store
import settings
import single_instance
import updater


class App:
    def __init__(self, root):
        self.root = root
        root.title(f"인스타 DM 매크로 v{config.APP_VERSION} (고객 {config.CUSTOMER_ID})")
        root.geometry("1000x660")

        self.driver = None          # browser 엔진일 때만 사용(selenium 드라이버)
        self.session = None         # 실행에 넘기는 세션(api=ApiSession, browser=드라이버)
        self.actions = None         # ig_api 또는 instagram_actions
        self.engine = None
        # 지금 붙어 있는 세션이 '어느 별명 / 어느 인스타 계정'인지.
        # v1.5.0: 이 값은 **막는 근거가 아니라 표시/기록용**이다. 실행 계정이 달라졌으면
        # 막지 않고 따라간다(= 고객이 크롬 창에서 인스타 자체 계정 전환을 쓴 정상 상황).
        self.session_label = None
        self.session_user_id = None
        self.session_username = None
        self.rows = []
        self.skipped_no_message = []
        self.logged_in = False
        # 로그인 흐름이 지금 돌고 있는가. [시작] 이 '로그인 안 됨' 으로 오판하지 않으려면
        # 이 상태를 알아야 한다(고객 실측: 로그인 확인이 28초 걸리는 동안 [시작] 을 눌렀다).
        self._login_busy = False
        self._start_pending = False

        self.account_var = tk.StringVar(value="default")
        self.live_var = tk.StringVar(value="현재 실행 계정: (로그인하면 표시됩니다)")
        self._label_by_display = {}
        self.engine_var = tk.StringVar(value=config.DEFAULT_ENGINE)
        self.username_var = tk.StringVar(value="")
        self.password_var = tk.StringVar(value="")
        self.excel_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="상태: 로그인 필요")
        self.stats_var = tk.StringVar(value="완료 0 / 팔로우 0 / DM 0 / 실패 0")
        self.cap_var = tk.StringVar(value=str(settings.get_daily_cap()))
        self.auto_switch_var = tk.BooleanVar(value=settings.get_auto_switch())
        # v1.7.0: 자동 업데이트가 막힌 것을 고객이 **볼 수 있어야** 한다. 고객 1898680 은
        # 두 번(1.3.1, 1.5.0) 옛 버전에 조용히 묶인 채로 계속 썼다. 이 줄은 항상 보이고,
        # 막히면 색이 바뀌며 수동 다운로드 주소를 같이 준다.
        self.update_var = tk.StringVar(
            value=f"프로그램 버전 v{config.APP_VERSION} (자동 업데이트 켜짐)")
        self._update_url = config.MANUAL_DOWNLOAD_URL
        self._update_reason = ""

        self._build_ui()
        self._migrate_legacy_bindings()
        self.updater_thread = updater.start_updater(
            stop_running_loop=self._stop_macro_silent, status_cb=self._log,
            blocked_cb=self._on_update_blocked, pre_swap_cb=self._quit_browser_for_swap)
        root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        try:
            if self.engine:
                self.engine.stop()
            if self.driver:
                self.driver.quit()
        except Exception:
            pass
        self.root.destroy()

    # ---------- UI ----------
    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        # 0) 버전/업데이트 줄. 평소엔 현재 버전만 조용히 보여 주고, 자동 교체가 막히면
        # 노란 경고로 바뀌며 [지금 업데이트] 와 다운로드 주소를 내민다.
        self.update_frame = tk.Frame(self.root, bg="#f4f6f8")
        self.update_frame.pack(fill="x", padx=8, pady=(8, 0))
        self.update_label = tk.Label(self.update_frame, textvariable=self.update_var,
                                     bg="#f4f6f8", fg="#333333", anchor="w",
                                     justify="left", wraplength=760)
        self.update_label.pack(side="left", padx=8, pady=6, fill="x", expand=True)
        ttk.Button(self.update_frame, text="지금 업데이트",
                   command=self.on_update_now_click).pack(side="right", padx=4, pady=4)
        ttk.Button(self.update_frame, text="다운로드 주소 복사",
                   command=self.on_copy_update_url_click).pack(side="right", padx=4, pady=4)

        acc_frame = ttk.LabelFrame(self.root, text="1) 인스타그램 계정")
        acc_frame.pack(fill="x", **pad)
        # v1.5.0: 별명을 매번 손으로 치면 'mugenboksa' / 'megenboksa' 처럼 한 글자 오타가
        # 조용히 **빈 프로필**을 하나 더 만든다(고객 실측). 그래서 저장된 별명은 목록에서
        # 고르게 하고, 새로 만들 때만 [+ 새 별명] 으로 확인을 거친다.
        ttk.Label(acc_frame, text="계정 별명(여러 계정 구분용):").grid(row=0, column=0, padx=6, pady=6, sticky="w")
        self.account_combo = ttk.Combobox(acc_frame, textvariable=self.account_var, width=26,
                                          state="readonly", values=["default"])
        self.account_combo.grid(row=0, column=1, padx=6, pady=6)
        self.account_combo.bind("<<ComboboxSelected>>", lambda _e: self._on_account_selected())
        ttk.Button(acc_frame, text="+ 새 별명", command=self.on_new_label_click).grid(row=0, column=2, padx=4)
        # 별명에 잘못 저장된 계정(예: 'mugenboksa · @mightysun_09')을 고객이 직접 지울 수 있어야
        # 한다. v1.5.0 에는 화면에 틀린 값이 보이는데 고칠 방법이 아예 없었다.
        ttk.Button(acc_frame, text="계정 기록 지우기",
                   command=self.on_forget_account_click).grid(row=0, column=3, padx=4)
        ttk.Radiobutton(acc_frame, text="빠른 방식(아이디/비번)", variable=self.engine_var,
                        value="api", command=self._on_engine_change).grid(row=0, column=4, padx=6, sticky="w")
        ttk.Radiobutton(acc_frame, text="크롬 창에서 직접 로그인", variable=self.engine_var,
                        value="browser", command=self._on_engine_change).grid(row=0, column=5, padx=6, sticky="w")

        self.cred_frame = ttk.Frame(acc_frame)
        self.cred_frame.grid(row=1, column=0, columnspan=6, sticky="w")
        ttk.Label(self.cred_frame, text="아이디:").grid(row=0, column=0, padx=6, pady=4, sticky="w")
        ttk.Entry(self.cred_frame, textvariable=self.username_var, width=22).grid(row=0, column=1, padx=6)
        ttk.Label(self.cred_frame, text="비밀번호:").grid(row=0, column=2, padx=6, sticky="w")
        ttk.Entry(self.cred_frame, textvariable=self.password_var, width=22, show="*").grid(row=0, column=3, padx=6)
        ttk.Label(self.cred_frame, text="(비밀번호는 저장하지 않습니다. 최초 1회만 입력)").grid(
            row=1, column=0, columnspan=4, padx=6, sticky="w")

        ttk.Button(acc_frame, text="로그인 / 계정 전환", command=self.on_login_click).grid(row=2, column=1, padx=6, pady=6)
        ttk.Button(acc_frame, text="다른 계정으로 로그인", command=self.on_switch_account_click).grid(row=2, column=2, padx=6)
        ttk.Button(acc_frame, text="로그아웃", command=self.on_logout_click).grid(row=2, column=3, padx=6)
        # 기본은 꺼짐. 켠 사람만 프로그램이 인스타 자체 '계정 전환'을 대신 눌러 준다.
        ttk.Checkbutton(acc_frame,
                        text="별명에 기억된 계정으로 크롬을 자동 전환 (권장: 끔)",
                        variable=self.auto_switch_var,
                        command=self._on_auto_switch_change).grid(row=2, column=4, columnspan=2,
                                                                  padx=6, sticky="w")
        ttk.Label(acc_frame, textvariable=self.status_var).grid(row=3, column=0, columnspan=6, padx=6, sticky="w")
        # 크롬 창에서 인스타 자체 계정 전환을 해도 되고, 그때 프로그램이 어느 계정으로 도는지는
        # 항상 여기에 그대로 보인다(막지 않는다).
        self.live_label = ttk.Label(acc_frame, textvariable=self.live_var, foreground="#0b5cad")
        self.live_label.grid(row=4, column=0, columnspan=6, padx=6, pady=(0, 6), sticky="w")
        ttk.Label(acc_frame,
                  text="크롬 창에서 직접 로그인한 계정으로 그대로 실행됩니다. "
                       "프로그램이 계정을 바꾸지 않습니다.",
                  foreground="#555555").grid(row=5, column=0, columnspan=6, padx=6,
                                             pady=(0, 6), sticky="w")
        self._on_engine_change()
        self._refresh_account_choices()

        file_frame = ttk.LabelFrame(self.root, text="2) 엑셀 파일 (C열=인스타 URL, F열=DM 문구)")
        file_frame.pack(fill="x", **pad)
        ttk.Entry(file_frame, textvariable=self.excel_var, width=70).grid(row=0, column=0, padx=6, pady=6)
        ttk.Button(file_frame, text="찾아보기", command=self.on_browse_click).grid(row=0, column=1, padx=6)
        ttk.Button(file_frame, text="불러오기", command=self.on_load_click).grid(row=0, column=2, padx=6)

        ctrl_frame = ttk.LabelFrame(self.root, text="3) 실행")
        ctrl_frame.pack(fill="x", **pad)
        ttk.Button(ctrl_frame, text="시작", command=self.on_start_click).grid(row=0, column=0, padx=6, pady=6)
        ttk.Button(ctrl_frame, text="중지", command=self.on_stop_click).grid(row=0, column=1, padx=6)
        ttk.Button(ctrl_frame, text="진행상황 초기화", command=self.on_reset_click).grid(row=0, column=2, padx=6)
        ttk.Label(ctrl_frame, textvariable=self.stats_var).grid(row=0, column=3, padx=16)
        ttk.Label(ctrl_frame, text="하루 최대 처리 인원:").grid(row=1, column=0, padx=6, pady=6, sticky="w")
        ttk.Spinbox(ctrl_frame, from_=config.DAILY_CAP_MIN, to=config.DAILY_CAP_MAX, width=6,
                    textvariable=self.cap_var, command=self._on_cap_change).grid(row=1, column=1, padx=6, sticky="w")
        ttk.Label(ctrl_frame,
                  text="계정 보호용 상한입니다. 이 인원까지 처리하면 그날은 자동으로 멈춥니다."
                  ).grid(row=1, column=2, columnspan=2, padx=6, sticky="w")

        log_frame = ttk.LabelFrame(self.root, text="로그")
        log_frame.pack(fill="both", expand=True, **pad)
        self.log_widget = scrolledtext.ScrolledText(log_frame, height=20, state="disabled")
        self.log_widget.pack(fill="both", expand=True, padx=6, pady=6)

    def _log(self, msg):
        def _do():
            self.log_widget.configure(state="normal")
            self.log_widget.insert("end", f"{time.strftime('%H:%M:%S')} {msg}\n")
            self.log_widget.see("end")
            self.log_widget.configure(state="disabled")
        try:
            self.root.after(0, _do)
        except Exception:
            print(msg)

    def _set_status(self, text):
        try:
            self.root.after(0, lambda: self.status_var.set(text))
        except Exception:
            pass

    # ---------- 자동 업데이트 (v1.7.0) ----------
    def _set_update_banner(self, text, warn=False, url=None):
        """버전 줄을 갱신한다. warn 이면 노란 경고로 바꿔 눈에 띄게 만든다."""
        if url:
            self._update_url = url

        def _do():
            try:
                self.update_var.set(text)
                bg = "#fff4d6" if warn else "#f4f6f8"
                fg = "#8a5a00" if warn else "#333333"
                self.update_frame.configure(bg=bg)
                self.update_label.configure(bg=bg, fg=fg)
            except Exception:
                pass
        try:
            self.root.after(0, _do)
        except Exception:
            pass

    def _on_update_blocked(self, latest, reason, url):
        """업데이터가 '자동 교체가 막혔다' 고 알려 올 때. 화면에서 절대 숨기지 않는다."""
        self._update_reason = reason or ""
        self._set_update_banner(
            f"자동 업데이트가 막혀 있습니다. 새 버전 v{latest} 이(가) 나왔지만 이 PC 에서 "
            f"교체되지 않았습니다(지금은 v{config.APP_VERSION}). [지금 업데이트] 를 누르거나 "
            f"아래 주소를 브라우저에 붙여 넣어 직접 받아 주세요:\n{url}",
            warn=True, url=url)
        self._log(f"[자동 업데이트] 막힘 사유: {reason}")

    def _quit_browser_for_swap(self):
        """exe 교체 직전에 크롬/드라이버를 정리한다.

        살아 있는 chromedriver 나 크롬 자식 프로세스가 프로그램 폴더를 붙잡고 있으면
        교체 스크립트의 rename/move 가 막힐 수 있다. 어차피 곧 재시작하므로 먼저 닫는다.
        """
        try:
            if self.engine:
                self.engine.stop()
        except Exception:
            pass
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass
        self.driver = None
        self.session = None

    def on_update_now_click(self):
        """[지금 업데이트] - 백오프를 무시하고 즉시 확인/교체를 시도한다."""
        self._log("[자동 업데이트] 지금 확인합니다...")
        self._set_update_banner("업데이트를 확인하는 중입니다...", warn=False)
        try:
            self.updater_thread.check_now(done_cb=self._on_update_now_result)
        except Exception as e:
            messagebox.showerror("자동 업데이트", f"업데이트 확인을 시작하지 못했습니다: {e}")

    def _on_update_now_result(self, result):
        result = result or {}
        status = result.get("status")
        detail = str(result.get("detail") or "")
        url = result.get("download_url") or self._update_url

        def _do():
            if status == "up_to_date":
                self._set_update_banner(
                    f"프로그램 버전 v{config.APP_VERSION} (최신입니다)", warn=False)
                messagebox.showinfo("자동 업데이트", detail or "이미 최신 버전입니다.")
                return
            if status == "busy":
                messagebox.showinfo("자동 업데이트", detail)
                return
            # blocked / failed / error / dev - 전부 사유를 그대로 보여 준다.
            self._update_reason = detail
            self._set_update_banner(
                f"자동 업데이트가 되지 않았습니다. 아래 주소에서 직접 받아 주세요:\n{url}",
                warn=True, url=url)
            self._log(f"[자동 업데이트] 실패 사유: {detail}")
            messagebox.showwarning(
                "자동 업데이트",
                f"이 PC 에서 자동 교체가 되지 않았습니다.\n\n"
                f"아래 주소를 브라우저에 붙여 넣어 새 파일을 받은 뒤,\n"
                f"압축을 풀어 지금 프로그램이 있는 폴더에 덮어써 주세요.\n\n{url}\n\n"
                f"[상세 사유]\n{detail[:900]}")
        try:
            self.root.after(0, _do)
        except Exception:
            pass

    def on_copy_update_url_click(self):
        url = self._update_url or config.MANUAL_DOWNLOAD_URL
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(url)
            self._log(f"[자동 업데이트] 다운로드 주소를 복사했습니다: {url}")
            messagebox.showinfo("자동 업데이트",
                                f"다운로드 주소를 복사했습니다.\n브라우저 주소창에 붙여 넣어 "
                                f"주세요.\n\n{url}")
        except Exception as e:
            messagebox.showerror("자동 업데이트", f"복사하지 못했습니다: {e}\n\n{url}")

    def _on_engine_change(self):
        """아이디/비번 입력칸은 빠른 방식(api)일 때만 보인다."""
        try:
            if self.engine_var.get() == "api":
                self.cred_frame.grid()
            else:
                self.cred_frame.grid_remove()
        except Exception:
            pass

    def _on_cap_change(self):
        try:
            self.cap_var.set(str(settings.set_daily_cap(self.cap_var.get())))
        except Exception:
            pass

    def _on_auto_switch_change(self):
        try:
            on = settings.set_auto_switch(self.auto_switch_var.get())
        except Exception:
            return
        self._log("크롬 자동 계정 전환을 " + ("켰습니다. 별명에 기억된 계정과 크롬 창의 계정이 "
                                            "다르면 프로그램이 인스타 계정 전환을 대신 누릅니다."
                                            if on else
                                            "껐습니다(권장). 크롬 창에 로그인된 계정 그대로 실행합니다."))
        bridge.remote_log("auto_switch_setting", f"on={on}", force=True)

    def _migrate_legacy_bindings(self):
        """v1.5.0 이전 로직이 만든 별명-계정 기록을 비운다. **조용히 지우지 않고 다 적는다.**

        고객 화면에는 `mugenboksa · @mightysun_09` 처럼 틀린 값이 떠 있었다. 그 값은 v1.4.0 이
        부모 세션의 계정을 그대로 별명에 묶어서 생긴 것이고, v1.5.0 은 그걸 정답으로 믿고
        브라우저를 끌고 갔다. 업그레이드 시점에 한 번 비워야 여섯 번째 판이 안 열린다.
        """
        try:
            import account_binding
            reset = account_binding.migrate_legacy()
        except Exception:
            return
        if not reset:
            return
        for item in reset:
            self._log(f"[업그레이드] 별명 '{item['label']}' 에 저장돼 있던 계정 기록"
                      f"(@{item['username'] or item['user_id'] or '?'})을 지웠습니다. "
                      f"이전 버전이 잘못 저장했을 수 있는 값이라 다음 로그인 때 실제 계정으로 "
                      f"다시 기억합니다.")
        try:
            bridge.remote_log(
                "binding_migration_v2",
                "reset=" + ";".join(f"{i['label']}->{i['username'] or '?'}/{i['user_id'] or '?'}"
                                    for i in reset), force=True)
        except Exception:
            pass
        self._refresh_account_choices()

    # ---------- 별명 목록(오타 방지) ----------
    def _current_label(self):
        """드롭다운에 보이는 표시 문자열('별명  ·  @아이디')에서 진짜 별명만 꺼낸다."""
        shown = (self.account_var.get() or "").strip()
        if shown in self._label_by_display:
            return self._label_by_display[shown]
        return shown.split("  ·  ", 1)[0].strip() or "default"

    def _progress_key_for(self, label):
        """진행상황을 셀 키. 계정 id 를 알면 계정 기준, 모르면 별명 기준."""
        try:
            import account_binding
            entry = account_binding.get(label) or {}
            return account_binding.run_key(entry.get("user_id"), label)
        except Exception:
            return label

    def _refresh_account_choices(self, select_label=None):
        """저장된 별명 + 각 별명이 지금 묶여 있는 인스타 계정을 드롭다운에 다시 채운다."""
        try:
            import account_binding
            entries = account_binding.entries()
        except Exception:
            entries = {}
        keep = select_label or self._current_label()
        labels = sorted(set(list(entries.keys()) + [keep or "default"]),
                        key=lambda s: s.lower())
        displays, mapping = [], {}
        for label in labels:
            entry = entries.get(label) or {}
            who = entry.get("username")
            uid = entry.get("user_id")
            if who:
                text = f"{label}  ·  @{who}"
            elif uid:
                text = f"{label}  ·  (id {uid})"
            elif entry.get("reset_from_username") or entry.get("reset_from_user_id"):
                text = f"{label}  ·  (계정 기록 지움 - 로그인하면 다시 기억)"
            else:
                text = f"{label}  ·  (아직 로그인 전)"
            displays.append(text)
            mapping[text] = label
        self._label_by_display = mapping
        def _apply():
            self.account_combo.configure(values=displays)
            for text, label in mapping.items():
                if label == keep:
                    self.account_var.set(text)
                    return
            if displays:
                self.account_var.set(displays[0])
        try:
            self.root.after(0, _apply)
        except Exception:
            _apply()

    def _on_account_selected(self):
        label = self._current_label()
        if self.session_label and self.session_label != label:
            self._log(f"'{label}' 별명의 크롬 창을 열려면 [로그인 / 계정 전환] 을 눌러 주세요. "
                      "(지금 그대로 [시작] 을 누르면 열려 있는 크롬 창에 로그인된 계정으로 "
                      "실행됩니다 - 프로그램이 계정을 바꾸지는 않습니다)")

    def on_forget_account_click(self):
        """이 별명에 저장된 '계정 기록'만 지운다. 로그인 세션/진행상황은 건드리지 않는다."""
        label = self._current_label()
        try:
            import account_binding
            entry = account_binding.get(label) or {}
        except Exception:
            entry = {}
        shown = entry.get("username") or entry.get("user_id")
        if not shown:
            messagebox.showinfo("안내", f"'{label}' 에는 저장된 계정 기록이 없습니다.")
            return
        if not messagebox.askyesno(
                "확인",
                f"별명 '{label}' 에 저장된 계정 기록(@{shown})을 지울까요?\n\n"
                "크롬 로그인과 진행상황은 그대로 두고, 프로그램이 기억하던 계정 이름만 지웁니다.\n"
                "다음에 [로그인 / 계정 전환] 또는 [시작] 을 하면 그때 실제로 로그인된 계정으로 "
                "다시 기억합니다."):
            return
        try:
            old = account_binding.forget_account(label)
        except Exception as e:
            messagebox.showerror("오류", f"계정 기록을 지우지 못했습니다: {e}")
            return
        self._log(f"별명 '{label}' 의 계정 기록(@{(old or {}).get('username') or shown})을 지웠습니다.")
        bridge.remote_log("binding_forgotten",
                          f"account={label} was={(old or {}).get('username')}/"
                          f"{(old or {}).get('user_id')}", force=True)
        self._refresh_account_choices(select_label=label)

    def on_new_label_click(self):
        """새 별명 추가 - 기존 별명과 한두 글자만 다른 오타면 먼저 되묻는다."""
        import difflib
        from tkinter import simpledialog
        try:
            import account_binding
            known = account_binding.labels()
        except Exception:
            known = []
        name = simpledialog.askstring("새 별명", "새로 만들 계정 별명을 입력해 주세요.",
                                      parent=self.root)
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in known:
            self._refresh_account_choices(select_label=name)
            return
        close = difflib.get_close_matches(name, known, n=1, cutoff=0.8)
        if close and not messagebox.askyesno(
                "확인", f"이미 '{close[0]}' 별명이 있습니다.\n"
                        f"'{name}' 을(를) 새로 만들면 로그인도 처음부터 다시 해야 합니다.\n\n"
                        f"그래도 새로 만들까요?\n(아니오 = '{close[0]}' 를 사용)"):
            self._refresh_account_choices(select_label=close[0])
            return
        self._refresh_account_choices(select_label=name)
        self._log(f"별명 '{name}' 을(를) 만들었습니다. [로그인 / 계정 전환] 을 눌러 로그인해 주세요.")

    def _set_live_account(self, ident, note=""):
        """지금 실제로 실행되는 계정을 화면에 표시한다(막는 대신 보여준다)."""
        ident = ident or {}
        who = ident.get("username")
        uid = ident.get("user_id")
        if who or uid:
            text = f"현재 실행 계정: @{who or '?'} (id {uid or '?'}, 확인방식 {ident.get('source')})"
        else:
            text = "현재 실행 계정: 확인하지 못했습니다"
        if note:
            text += f" - {note}"
        try:
            self.root.after(0, lambda: self.live_var.set(text))
        except Exception:
            self.live_var.set(text)

    # ---------- 계정 ----------
    def on_login_click(self):
        label = self._current_label()
        if self.engine_var.get() == "api":
            threading.Thread(target=self._api_login_flow, args=(label,), daemon=True).start()
        else:
            threading.Thread(target=self._login_flow, args=(label,), daemon=True).start()

    def on_switch_account_click(self):
        """이 별명의 저장된 로그인을 지우고 처음부터 로그인한다.

        같은 별명에 다른 인스타 계정을 쓰고 싶을 때(또는 이전 계정으로 고정돼 버렸을 때) 쓴다.
        프로필 폴더를 통째로 지우므로 반드시 새 로그인 화면이 뜬다.
        """
        label = self._current_label()
        if self.engine_var.get() == "api":
            messagebox.showinfo("안내", "'크롬 창에서 직접 로그인' 방식에서만 사용할 수 있습니다.")
            return
        if not messagebox.askyesno(
                "확인", f"'{label}' 별명의 저장된 로그인을 지우고 새로 로그인할까요?\n"
                        "(진행상황은 지워지지 않습니다)"):
            return
        threading.Thread(target=self._switch_account_flow, args=(label,), daemon=True).start()

    def _switch_account_flow(self, label):
        import shutil
        if self.driver is not None:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None
        self.session = None
        self.actions = None
        self.logged_in = False
        self.session_label = None
        self.session_user_id = None
        try:
            import account_binding
            account_binding.unbind(label)
        except Exception:
            pass
        profile_dir = config.profile_dir_for(label)
        try:
            shutil.rmtree(profile_dir, ignore_errors=True)
            self._log(f"'{label}' 의 저장된 로그인을 지웠습니다. 새로 로그인해 주세요.")
        except Exception as e:
            self._log(f"로그인 정보 삭제 중 오류: {e}")
        bridge.remote_log("account_reset", f"account={label}", force=True)
        self._login_flow(label)

    def _api_login_flow(self, label, verification_code=None):
        """instagrapi 로그인. 저장된 세션이 있으면 비밀번호 없이 붙는다."""
        import ig_api
        self._set_status("상태: 접속 중...")
        try:
            session = ig_api.load_session(label, proxy=config.API_PROXY, log=self._log)
            if session is None:
                username = self.username_var.get().strip()
                password = self.password_var.get()
                if not username or not password:
                    self._log("저장된 세션이 없습니다. 아이디와 비밀번호를 입력한 뒤 다시 눌러주세요.")
                    self._set_status("상태: 아이디/비밀번호 입력 필요")
                    return
                session = ig_api.login(label, username, password,
                                       verification_code=verification_code,
                                       proxy=config.API_PROXY, log=self._log)
        except ig_api.LoginNeedsCode as need:
            self._ask_code_and_retry(label, need)
            return
        except Exception as e:
            self._log(f"로그인 실패: {e}")
            self._set_status("상태: 로그인 실패")
            bridge.remote_log("api_login_failed", str(e)[:500], force=True)
            return

        self.session = session
        self.actions = ig_api
        self.logged_in = True
        self.session_label = label
        self.session_user_id = None
        self.password_var.set("")  # 화면에도 비밀번호를 남기지 않는다
        who = ig_api.account_username(session)
        self._log(f"'{label}' 접속 완료{f' (@{who})' if who else ''}.")
        self._set_status(f"상태: 로그인됨 ({label}{f' / @{who}' if who else ''})")
        bridge.remote_log("api_login_ok", f"account={label}", force=True)

    def _ask_code_and_retry(self, label, need):
        """2단계 인증/본인확인 코드를 받아 같은 흐름을 다시 탄다(코드는 저장하지 않는다)."""
        from tkinter import simpledialog

        def _prompt():
            code = simpledialog.askstring("인증 코드", f"{need}\n\n받은 6자리 코드를 입력해 주세요.",
                                          parent=self.root)
            if not code:
                self._log("인증 코드 입력이 취소됐습니다.")
                self._set_status("상태: 인증 코드 필요")
                return
            threading.Thread(target=self._api_login_flow, args=(label, code.strip()),
                             daemon=True).start()

        self.root.after(0, _prompt)

    def _login_flow(self, label):
        """로그인 흐름 한 겹 감싸기. 두 가지를 보장한다.

        1) `_login_busy` 가 무슨 일이 있어도 내려간다. 안 그러면 [시작] 이 'pending' 으로 굳어
           3분 기다렸다가 실패한다.
        2) 예외가 나도 **조용히 죽지 않는다.** v1.5.0 까지는 이 스레드가 예외로 사라지면 화면에
           아무 말도 없이 로그인 미완료 상태로 남았고, 고객은 [시작] 에서만 그 사실을 알았다.
           이제는 로그에 남기고, [시작] 은 살아 있는 크롬 창을 직접 보고 판단한다.
        """
        self._login_busy = True
        try:
            self._login_flow_inner(label)
        except Exception as e:
            self._log(f"로그인 처리 중 오류: {e}")
            self._set_status("상태: 로그인 확인 실패")
            try:
                bridge.remote_log("login_flow_error", f"account={label} err={str(e)[:300]}",
                                  force=True)
            except Exception:
                pass
        finally:
            self._login_busy = False

    def _login_flow_inner(self, label):
        import browser
        self._log(f"'{label}' 계정 프로필로 크롬을 엽니다...")
        self._set_status("상태: 브라우저 준비 중...")
        try:
            if self.driver is not None:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                self.driver = None
            profile_dir = config.profile_dir_for(label)
            os.makedirs(profile_dir, exist_ok=True)
            self.driver = browser.build_driver(profile_dir, log=self._log)
        except Exception as e:
            self._log(f"브라우저 시작 실패: {e}")
            self._set_status("상태: 브라우저 시작 실패")
            return

        import account_binding
        import instagram_actions as ig
        if ig.is_logged_in(self.driver):
            # **먼저** 실행 가능 상태로 만든다. v1.5.0 은 계정 판독/전환(고객 실측 28초)이 전부
            # 끝난 뒤에야 이 값들을 채웠고, 그 사이에 [시작] 을 누른 고객은 "먼저 계정 로그인을
            # 완료해 주세요" 팝업을 맞았다. 로그인된 창이 이미 있으면 그 순간부터 실행 가능이다.
            self.session = self.driver
            self.actions = ig
            self.logged_in = True
            self.session_label = label

            ident = ig.resolve_identity(self.driver)
            uid, who = ident.get("user_id"), ident.get("username")
            # check() 는 기록을 살아 있는 계정으로 **고쳐 쓴다.** 그러니 '체크박스를 켠 사람만'
            # 쓰는 자동 전환이 무엇을 원했는지는 고치기 전에 미리 꺼내 둬야 한다.
            # (이걸 안 하면 자동 전환은 언제나 want == 현재계정 이 되어 조용히 죽는다)
            bound_before = account_binding.get(label) or {}
            verdict, detail = account_binding.check(label, uid, who)
            bridge.remote_log(
                "login_identity",
                f"account={label} user={who} uid={uid} src={ident.get('source')} verdict={verdict}",
                force=True)
            if verdict == "unknown":
                # 계정 id 자체를 못 읽었다 = 사실상 로그인 상태가 아니다. 새로 로그인 받는다.
                self.logged_in = False
                self.session = None
                self._log("로그인 상태를 확인하지 못했습니다. 새로 로그인해 주세요.")
                self._wait_manual_login(label, ig, account_binding)
                return

            if verdict == "rebound":
                # v1.5.0 은 여기서 **저장된 계정으로 브라우저를 되돌렸다**. 고객이 직접 로그인한
                # 서브계정에서 부모 계정으로 끌려 나갔다(2026-08-04 실측, 고객이 눈으로 확인).
                # v1.6.0: 살아 있는 계정이 이긴다. 브라우저는 건드리지 않고 기억만 고친다.
                self._log(f"[안내] {detail}")
                bridge.remote_log("login_binding_rewritten",
                                  f"account={label} live_uid={uid} live_user={who} "
                                  f"src={ident.get('source')}", force=True)
            if self._maybe_auto_switch(label, "login", ig,
                                       want=bound_before.get("username")):
                ident = ig.resolve_identity(self.driver)
                uid, who = ident.get("user_id"), ident.get("username")

            self.session_user_id = uid
            self.session_username = who
            account_binding.bind(label, uid, who)
            self._set_live_account(ident)
            self._refresh_account_choices(select_label=label)
            self._log(f"'{label}' 프로필에 이미 로그인되어 있습니다"
                      f"{f' (@{who})' if who else ''}. 이 계정 그대로 실행됩니다.")
            self._set_status(f"상태: 로그인됨 ({label}{f' / @{who}' if who else ''})")
            bridge.remote_log("login_reused",
                              f"account={label} user={who} uid={uid} src={ident.get('source')}",
                              force=True)
            return

        ig.goto_login_screen(self.driver)
        self._wait_manual_login(label, ig, account_binding)

    def _maybe_auto_switch(self, label, where, ig, want=None):
        """체크박스를 **켠 경우에만** 인스타 자체 계정 전환을 대신 눌러 준다(기본 꺼짐).

        기본이 꺼짐인 이유는 정직하게: 이 기능이 지금까지 실제로 증명한 효과는 고객을 자기가
        로그인한 계정에서 끌어낸 것 하나뿐이다(2026-08-04 `login_switch_attempt ok=True` ->
        고객: "크롬창에서는 강제로 A계정의 B아이디에서 A아이디로 변경되더라구요").

        `want` 는 호출자가 **기록을 고쳐 쓰기 전에** 읽어 둔 별명의 저장 계정. 안 주면 지금
        저장된 값을 읽는다. 반환값은 '실제로 전환에 성공했는가'.
        """
        try:
            if not settings.get_auto_switch():
                return False
        except Exception:
            return False
        if want is None:
            try:
                import account_binding
                want = (account_binding.get(label) or {}).get("username")
            except Exception:
                return False
        driver = self.session or self.driver
        if driver is None:
            return False
        try:
            cur = ig.resolve_identity(driver)
        except Exception:
            return False
        if not want or (cur.get("username") or "").lower() == str(want).lower():
            return False
        self._log(f"[자동 전환 켜짐] '{label}' 은 @{want} 로 기억돼 있어 인스타 계정 전환을 시도합니다...")
        ok, detail = ig.switch_to_account(driver, want, log=self._log)
        bridge.remote_log(f"{where}_switch_attempt",
                          f"account={label} want={want} ok={ok} detail={detail[:300]}", force=True)
        self._log(f"[계정 전환] {detail}")
        if not ok:
            self._switch_failure_dump(label, detail)
            self._log("[안내] 전환하지 못했습니다. 크롬 창에 로그인된 계정 그대로 진행합니다.")
        return ok

    def _switch_failure_dump(self, label, detail):
        """계정 전환 UI 를 못 찾았을 때 화면 DOM 을 통째로 올린다.

        인스타 전환 UI 는 라이브 계정 없이는 셀렉터를 확정할 수 없다. 실패한 그 순간의 DOM 이
        다음 판에 셀렉터를 못박을 유일한 증거이므로, 실패를 조용히 넘기지 않고 올린다.
        """
        try:
            import diag_collector
            path = diag_collector.capture_zip(self.driver, f"switch_fail_{label}", detail[:1500])
            bridge.upload_run(f"account_switch_failed account={label} detail={detail[:500]}",
                              zip_path=path, kind="switchdiag")
        except Exception:
            pass

    def _wait_manual_login(self, label, ig, account_binding):
        """크롬 창에서 사람이 직접 로그인할 때까지 기다리고, 성공하면 별명에 계정을 묶는다."""
        self._log("크롬 창에서 인스타그램 아이디/비밀번호를 직접 입력해 로그인해 주세요. "
                  "(자동입력 안 함 - 2단계 인증/보안 확인도 그대로 통과 가능)")
        self._set_status("상태: 로그인 대기 중 (창에서 직접 로그인)")
        ok = ig.wait_for_manual_login(
            self.driver, timeout_s=600,
            poll_cb=lambda: self._set_status("상태: 로그인 대기 중..."))
        if ok:
            self.logged_in = True
            self.session = self.driver
            self.actions = ig
            self.session_label = label
            ident = ig.resolve_identity(self.driver)
            uid, who = ident.get("user_id"), ident.get("username")
            self.session_user_id = uid
            self.session_username = who
            # 이 별명은 지금 로그인한 계정으로 (다시) 묶는다. 다음 실행부터 이 값으로 대조한다.
            account_binding.bind(label, uid, who)
            self._set_live_account(ident)
            self._refresh_account_choices(select_label=label)
            self._log(f"'{label}' 로그인 확인됨{f' (@{who})' if who else ''}.")
            self._set_status(f"상태: 로그인됨 ({label}{f' / @{who}' if who else ''})")
            bridge.remote_log("login_ok",
                              f"account={label} user={who} uid={uid} src={ident.get('source')}",
                              force=True)
        else:
            self._log("로그인 대기 시간이 초과됐습니다. 다시 시도해 주세요.")
            self._set_status("상태: 로그인 대기 시간 초과")
            bridge.remote_log("login_timeout", f"account={label}", force=True)

    def on_logout_click(self):
        if self.session is None:
            messagebox.showinfo("안내", "먼저 로그인/계정 전환을 눌러주세요.")
            return
        threading.Thread(target=self._logout_flow, daemon=True).start()

    def _logout_flow(self):
        try:
            self.actions.logout(self.session, log=self._log)
        except Exception as e:
            self._log(f"로그아웃 처리 중 오류: {e}")
        self.logged_in = False
        self.session = None
        self.session_label = None
        self.session_user_id = None
        self._set_status("상태: 로그아웃됨")
        bridge.remote_log("logout", f"account={self._current_label()}", force=True)

    # ---------- 엑셀 ----------
    def on_browse_click(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.excel_var.set(path)

    def on_load_click(self):
        path = self.excel_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("오류", "엑셀 파일을 선택해 주세요.")
            return
        try:
            self.rows, self.skipped_no_message = excel_reader.load_rows(path)
        except Exception as e:
            messagebox.showerror("오류", f"엑셀을 읽는 중 오류: {e}")
            return
        label = self._current_label()
        done = progress_store.load_done_rows(
            path, self._progress_key_for(label)) | progress_store.load_done_rows(path, label)
        self._log(f"엑셀 로드 완료: 처리 대상 {len(self.rows)}행 "
                   f"(이미 완료 {len([r for r in self.rows if r.row_no in done])}행), "
                   f"F열 비어 스킵 {len(self.skipped_no_message)}행")

    # ---------- 실행 ----------
    # [시작] 이 '로그인 됐나'를 판단하는 기준은 **살아 있는 크롬 창**이지 self.logged_in 플래그가
    # 아니다. v1.5.0 은 플래그 하나만 봤고, 그 플래그는 로그인 흐름이 계정 판독/전환을 다 끝낸
    # 뒤에야 켜졌다. 고객 실측(2026-08-04):
    #     09:07:34 크롬 엽니다 ... 09:07:45 엑셀 로드 ... 09:08:02 로그인되어 있습니다
    # 그 사이(약 28초)에 [시작] 을 누르면 화면에는 '로그인됨'이 떠 있는데도 "먼저 계정 로그인을
    # 완료해 주세요" 팝업이 났다. 아래 세 상태를 구분해 각각 다르게 처리한다.
    #   ready   - 지금 바로 실행 가능
    #   pending - 로그인 흐름이 아직 돌고 있다 -> **팝업 대신 기다렸다가 자동 시작**
    #   none    - 크롬 창 자체가 없다/로그인 안 됨 -> 그때만 안내 팝업
    START_WAIT_POLL_MS = 1000
    START_WAIT_MAX_POLLS = 180          # 최대 3분까지 기다렸다가 자동 시작

    def _start_readiness(self):
        """(상태, 설명). 플래그가 아니라 지금 살아 있는 상태에서 유도한다."""
        if self.session is not None and self.logged_in:
            return "ready", "session"
        if self.engine_var.get() == "api":
            # 빠른 방식은 브라우저가 없어 창을 들여다볼 수 없다. 진행 중이면 기다린다.
            return ("pending", "api_login") if self._login_busy else ("none", "api_no_session")
        if self.driver is None:
            return ("pending", "browser_starting") if self._login_busy else ("none", "no_browser")
        if self._login_busy:
            # 로그인 스레드가 같은 드라이버로 명령을 보내는 중이다. 끼어들지 말고 기다린다.
            return "pending", "login_running"
        import instagram_actions as ig
        try:
            live = ig.session_is_live(self.driver)
        except Exception:
            # 창을 들여다볼 수 없다(창이 죽었다). 그건 '로그인 안 됨' 이 맞다.
            return "none", "window_unreadable"
        if not live:
            return "none", "window_not_logged_in"

        # 창은 로그인돼 있는데 플래그만 안 켜져 있다(로그인 스레드가 중간에 죽었거나 예외로
        # 빠져나간 경우). 플래그를 창 기준으로 맞추고 그대로 실행한다.
        # 아래 부수 작업(로그/라벨)이 실패해도 실행 가능 판정은 취소되지 않는다. 여기를 통째로
        # try 로 감싸면 사소한 예외가 다시 '로그인하세요' 팝업으로 둔갑한다.
        self.session = self.driver
        self.actions = ig
        self.logged_in = True
        try:
            self.session_label = self._current_label()
            self._log("[안내] 크롬 창이 이미 로그인 상태라 그대로 실행합니다.")
            bridge.remote_log("start_session_adopted",
                              f"account={self.session_label}", force=True)
        except Exception:
            pass
        return "ready", "adopted_live_window"

    def on_start_click(self):
        if self.engine is not None and self.engine.is_alive():
            messagebox.showinfo("안내", "이미 실행 중입니다.")
            return
        self._start_pending = True
        self._attempt_start(0)

    def _attempt_start(self, polls):
        if not self._start_pending:
            return
        state, why = self._start_readiness()
        if state == "pending":
            if polls == 0:
                self._log("로그인 확인이 끝나는 대로 자동으로 시작합니다. 잠시만 기다려 주세요.")
            if polls < self.START_WAIT_MAX_POLLS:
                self._set_status("상태: 로그인 확인 중... (끝나면 자동으로 시작합니다)")
                try:
                    self.root.after(self.START_WAIT_POLL_MS,
                                    lambda: self._attempt_start(polls + 1))
                    return
                except Exception:
                    pass
            self._start_pending = False
            bridge.remote_log("start_wait_timeout", f"why={why} polls={polls}", force=True)
            messagebox.showerror("오류", "로그인 확인이 끝나지 않았습니다. "
                                         "[로그인 / 계정 전환] 을 다시 눌러 주세요.")
            return
        if state == "none":
            self._start_pending = False
            bridge.remote_log("start_no_session", f"why={why}", force=True)
            messagebox.showerror("오류", "먼저 [로그인 / 계정 전환] 을 눌러 크롬 창에서 "
                                         "인스타그램에 로그인해 주세요.")
            return
        self._start_pending = False
        self._begin_run()

    def _begin_run(self):
        if not self.rows:
            self.on_load_click()
        if not self.rows:
            messagebox.showerror("오류", "처리할 행이 없습니다. 엑셀을 확인해 주세요.")
            return
        import macro_engine
        label = self._current_label()
        run_key = self._resolve_run_account(label)
        self.engine = macro_engine.MacroEngine(
            self.session, self.rows, self.excel_var.get().strip(), run_key,
            log_cb=self._log, done_cb=self._on_row_done,
            daily_cap=settings.get_daily_cap(), halt_cb=self._on_halt,
            actions=self.actions)
        self.engine.start()
        self._log("매크로를 시작합니다.")

    def _resolve_run_account(self, label):
        """[시작] 직전: **지금 크롬 창이 실제로 어느 계정으로 동작하는지** 확인하고 그 계정으로
        돌린다. 진행상황/하루 상한을 셀 키(run_key)를 돌려준다.

        v1.4.0 은 여기서 '시작할 때와 계정이 달라졌다'며 [시작] 을 아예 막았다. 그런데 이 고객은
        부모 계정 하나에 서브계정이 여러 개 붙어 있어서, 크롬 창에서 인스타 자체 계정 전환을
        쓰는 것이 정상 사용법이다. 그 정상 동작이 전부 차단으로 이어져 프로그램을 쓸 수 없었다
        (실측: start_blocked_uid_drift 3회 반복 + 재로그인 루프).

        v1.5.0 은 '막지 않는다' 는 지켰지만 여기서 **브라우저를 저장값 쪽으로 끌고 갔다.**
        고객이 눈으로 본 그대로: "크롬창에서는 강제로 A계정의 B아이디에서 A아이디로 변경되더라구요".
        고객이 직접 로그인해 둔 서브계정에서 [시작] 을 누르는 순간 부모 계정으로 튕겨 나간 것이다.

        v1.6.0 원칙:
          - **절대 막지 않는다.** 팔로우/DM 이 실행되는 그 계정이 정답이고, 프로그램은 따라간다.
          - **절대 브라우저 계정을 바꾸지 않는다.** 저장값이 다르면 저장값을 고쳐 쓴다.
            (체크박스로 자동 전환을 켠 사람만 예외 - 기본은 꺼짐)
          - 무엇으로 도는지는 화면(현재 실행 계정) + 진단 로그에 항상 남긴다.
          - 진행상황/하루 상한은 별명이 아니라 **실제 계정 id** 로 센다(오타 별명 2개가 같은
            계정을 가리켜도 이미 DM 보낸 사람에게 다시 보내지 않는다).
        """
        run_key = label
        if self.engine_var.get() == "api":
            return run_key
        try:
            import account_binding
            import instagram_actions as ig
        except Exception:
            return run_key
        try:
            ident = ig.resolve_identity(self.session)
        except Exception:
            ident = {"user_id": None, "username": None, "source": "none"}
        uid, who = ident.get("user_id"), ident.get("username")

        if not uid:
            self._set_live_account(ident, "확인 실패 - 그대로 진행합니다")
            self._log("[안내] 지금 로그인된 계정을 확인하지 못했지만 그대로 시작합니다.")
            bridge.remote_log("start_identity_unknown",
                              f"account={label} src={ident.get('source')}", force=True)
            return run_key

        bound = account_binding.get(label) or {}
        want_id, want_user = bound.get("user_id"), bound.get("username")
        if want_id and str(want_id) != str(uid):
            # 여기가 고객을 다섯 번째로 막은 자리다. v1.5.0 은 이 지점에서 switch_to_account 를
            # 불러 크롬을 저장값 계정으로 되돌렸다. v1.6.0 은 저장값을 고친다.
            self._log(f"[안내] '{label}' 에 기억돼 있던 계정은 @{want_user or want_id} 였는데 "
                      f"크롬 창은 @{who or uid} 로 로그인돼 있습니다. "
                      f"지금 로그인된 계정(@{who or uid})으로 실행하고, 별명 기억을 그 계정으로 "
                      f"고쳐 둡니다. 크롬 창의 계정은 바꾸지 않습니다.")
            bridge.remote_log("start_binding_rewritten",
                              f"account={label} bound={want_id}/{want_user} live={uid}/{who} "
                              f"src={ident.get('source')}", force=True)

        # 체크박스를 켠 사람만: 저장값 계정으로 인스타 자체 전환을 대신 눌러 준다(기본 꺼짐).
        if self._maybe_auto_switch(label, "start", ig, want=want_user):
            ident = ig.resolve_identity(self.session)
            uid, who = ident.get("user_id"), ident.get("username")

        # 별명은 항상 '지금 실제로 도는 계정' 으로 다시 묶는다(다음에 열었을 때 그대로 보이게).
        account_binding.bind(label, uid, who)
        self.session_user_id = uid
        self.session_username = who
        self.session_label = label
        self._set_live_account(ident)
        self._refresh_account_choices(select_label=label)

        run_key = account_binding.run_key(uid, label)
        if run_key != label:
            # 예전 버전은 별명으로 진행상황을 저장했다. 계정 키로 옮길 때 승계하지 않으면
            # 이미 DM 을 보낸 사람에게 한 번 더 보낸다.
            try:
                path = self.excel_var.get().strip()
                if progress_store.migrate(path, label, run_key):
                    self._log(f"'{label}' 의 진행상황을 계정(@{who or uid}) 기준으로 승계했습니다.")
                progress_store.migrate_daily(label, run_key)
            except Exception:
                pass
        bridge.remote_log("start_identity",
                          f"account={label} run_key={run_key} user={who} uid={uid} "
                          f"src={ident.get('source')}", force=True)
        # 세 소스를 한 줄로 같이 남긴다: 다음 실행 로그만 보고 '소스별로 값이 갈리는지'를
        # 추측이 아니라 실측으로 판정하기 위해서다.
        try:
            bridge.remote_log("identity_sources",
                              f"account={label} " + ig.identity_report(self.session), force=True)
        except Exception:
            pass
        return run_key

    def _on_halt(self, reason, message):
        """엔진이 스스로 멈췄을 때: 로그만으로는 놓치기 쉬우니 팝업으로 확실히 알린다."""
        def _show():
            self.status_var.set(f"상태: 중단됨 ({reason.split(':', 1)[0]})")
            messagebox.showwarning("매크로 중단", message)
        try:
            self.root.after(0, _show)
        except Exception:
            pass

    def _on_row_done(self, row_no, follow_ok, dm_ok):
        if self.engine:
            s = self.engine.stats
            processed = s["dm_sent"] + s["failed"]
            self.root.after(0, lambda: self.stats_var.set(
                f"완료 {processed} / 팔로우 {s['followed']} / DM {s['dm_sent']} / 실패 {s['failed']}"))

    def on_stop_click(self):
        if self.engine:
            self.engine.stop()
            self._log("중지를 요청했습니다 (진행 중인 사람까지만 마치고 멈춥니다).")

    def _stop_macro_silent(self):
        if self.engine:
            self.engine.stop()

    def on_reset_click(self):
        path = self.excel_var.get().strip()
        label = self._current_label()
        if not path:
            messagebox.showinfo("안내", "엑셀 파일을 먼저 선택해 주세요.")
            return
        if messagebox.askyesno("확인", "이 엑셀+계정 조합의 진행상황을 초기화할까요? "
                                       "(처음부터 다시 팔로우+DM 을 시도합니다)"):
            progress_store.reset(path, label)
            progress_store.reset(path, self._progress_key_for(label))
            self._log("진행상황을 초기화했습니다.")


# ---------------- 데모/자동테스트 모드 ----------------

def _make_demo_excel():
    import openpyxl
    fd_path = os.path.join(os.environ.get("TEMP", "."), "insta_dm_macro_demo.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "번호"; ws["B1"] = "이름"; ws["C1"] = "인스타그램URL"
    ws["D1"] = "기타1"; ws["E1"] = "기타2"; ws["F1"] = "DM문구"; ws["G1"] = "섭외메시지 캡처"
    demo_rows = [
        (1, "데모유저1", "https://www.instagram.com/demo_user_one/", "-", "-",
         "안녕하세요! 협업 제안드리고 싶어 연락드려요 :)", "(무시됨)"),
        (2, "데모유저2", "https://www.instagram.com/demo_user_two/", "-", "-",
         "반갑습니다! 프로필 잘 보고 DM 드립니다.", "(무시됨)"),
    ]
    for i, r in enumerate(demo_rows, start=2):
        ws.cell(row=i, column=1, value=r[0])
        ws.cell(row=i, column=2, value=r[1])
        ws.cell(row=i, column=3, value=r[2])
        ws.cell(row=i, column=4, value=r[3])
        ws.cell(row=i, column=5, value=r[4])
        ws.cell(row=i, column=6, value=r[5])
        ws.cell(row=i, column=7, value=r[6])
    wb.save(fd_path)
    return fd_path


def _run_guidemo(root, app):
    """실제 크롬/인스타그램 접속 없이: 엑셀 파싱 -> 팔로우/DM 순서 -> 랜덤 대기 계산까지
    코드 경로를 실제로 실행하고 로그창/통계에 실측 결과를 남긴다(CI 스크린샷 증거용).
    """
    demo_path = _make_demo_excel()
    app.excel_var.set(demo_path)
    app.account_var.set("demo_account")
    rows, skipped = excel_reader.load_rows(demo_path)
    app.rows = rows
    app._log(f"[데모] 엑셀 파싱 완료: {len(rows)}행 (C/F열만 읽음, 무관한 컬럼은 무시)")
    app.status_var.set("상태: 로그인됨 (demo_account) [데모 모드]")

    followed = dm_sent = 0
    for row in rows:
        app._log(f"[데모] {row.row_no}행 @{row.username} 프로필 접속 -> 팔로우 버튼 클릭")
        followed += 1
        wait_s = random.uniform(config.DELAY_AFTER_FOLLOW_MIN, config.DELAY_AFTER_FOLLOW_MAX)
        app._log(f"[데모] 팔로우 직후 랜덤 대기 {wait_s:.1f}초 (탐지 회피)")
        app._log(f"[데모] {row.row_no}행 @{row.username} 에게 DM 발송: \"{row.message[:30]}\"")
        dm_sent += 1
        gap_s = random.uniform(config.DELAY_BETWEEN_PEOPLE_MIN, config.DELAY_BETWEEN_PEOPLE_MAX)
        app._log(f"[데모] 다음 사람까지 랜덤 대기 {gap_s:.1f}초 (계정 보호)")
    app.stats_var.set(f"완료 {len(rows)} / 팔로우 {followed} / DM {dm_sent} / 실패 0")
    app._log(f"[데모] 완료: F열 없어 스킵 {len(skipped)}행")

    # v1.6.0 에서 고친 것들을 이 창에서 **실제 코드로** 실행해 보여준다(CI 스크린샷 증거).
    # 고객이 다섯 번 리포트한 두 가지를 그대로 재현한다: (1) [시작] 이 로그인을 못 알아본다,
    # (2) 별명에 잘못 저장된 계정이 크롬 창을 다른 계정으로 끌고 간다.
    try:
        import account_binding
        import instagram_actions as ig
        import updater

        class _DriftDriver:
            """로그인 때와 [시작] 때 계정이 달라진 상황을 그대로 재현한다(고객 실측 사고)."""
            def __init__(self, uid, name):
                self.uid, self.name = uid, name
            current_url = "https://www.instagram.com/"
            def execute_async_script(self, *_a, **_k):
                return {"id": self.uid, "username": self.name}
            def execute_script(self, *_a, **_k):
                return None
            def get_cookies(self):
                return [{"name": "sessionid", "value": "demo"},
                        {"name": "ds_user_id", "value": "67584782851"}]
            def find_elements(self, *_a, **_k):
                return []

        app._log(f"[v1.6.0 확인] 팔로우 버튼 판정: "
                 f"'맞팔로우'->{ig.classify_follow_text('맞팔로우')}, "
                 f"'팔로잉'->{ig.classify_follow_text('팔로잉')}, "
                 f"'팔로우 취소'->{ig.classify_follow_text('팔로우 취소')}")
        drift = _DriftDriver("42105781019", "mugenboksa")
        ident = ig.resolve_identity(drift)
        app._log(f"[v1.6.0 확인] 실행 계정 판독(단일 소스): {ig.identity_str(ident)} "
                 f"- 쿠키(ds_user_id=67584782851)가 아니라 실제 실행 계정을 따른다")
        app._set_live_account(ident)

        # (1) [시작] 준비 판정: 플래그가 아니라 살아 있는 크롬 창을 보고 판단한다.
        saved = (app.driver, app.session, app.logged_in)
        app.driver, app.session, app.logged_in = drift, None, False
        state, why = app._start_readiness()
        app.driver, app.session, app.logged_in = saved
        app._log(f"[v1.6.0 확인] 로그인 플래그가 꺼져 있어도 크롬 창이 로그인 상태면 "
                 f"[시작] 준비 판정 = {state}({why}) - v1.5.0 은 여기서 "
                 f"'먼저 계정 로그인을 완료해 주세요' 팝업을 띄웠다")

        # (2) 오염된 바인딩: 살아 있는 계정이 이기고 저장값이 고쳐진다(브라우저는 그대로).
        account_binding.bind("__demo__", "67584782851", "mightysun_09")   # 고객의 오염된 기록
        verdict, detail = account_binding.check("__demo__", "42105781019", "mugenboksa")
        app._log(f"[v1.6.0 확인] 저장값(@mightysun_09) vs 크롬 창(@mugenboksa) -> "
                 f"판정 '{verdict}': 저장값이 @"
                 f"{account_binding.get('__demo__')['username']} 로 고쳐졌다. "
                 f"크롬 창의 계정은 바꾸지 않는다(자동 전환 기본값 = "
                 f"{'켬' if settings.get_auto_switch() else '끔'})")
        app._log(f"[v1.6.0 확인] 진행상황 키: 별명이 아니라 계정 기준 "
                 f"'{account_binding.run_key('42105781019', '__demo__')}' "
                 f"(오타 별명 2개가 같은 계정이어도 중복 DM 안 감)")
        forgotten = account_binding.forget_account("__demo__")
        app._log(f"[v1.6.0 확인] [계정 기록 지우기]: @{forgotten['username']} 기록을 지우고 "
                 f"별명은 남긴다(고객이 화면에서 직접 고칠 수 있다)")
        account_binding.unbind("__demo__")
        app._log(f"[v1.6.0 확인] 자동 업데이트 대상 경로(sys.executable): {updater.target_exe_path()}")

        # ---- v1.7.0: '사유 미기록' 이 사라졌다는 것을 실제 코드로 보여 준다 ----
        # 고객의 진짜 마커를 건드리지 않도록 임시 폴더로 갈아 끼우고, 데모 진단이 서버로
        # 올라가지 않게 리포터도 잠시 끈다.
        import tempfile as _tf
        demo_dir = _tf.mkdtemp(prefix="upd_demo_")
        keep = (updater._STATE_PATH, updater._TRAIL_PATH, updater.remote_log)
        try:
            updater._STATE_PATH = os.path.join(demo_dir, "update_state.json")
            updater._TRAIL_PATH = os.path.join(demo_dir, "update_trail.log")
            updater.remote_log = lambda *_a, **_k: None

            # 고객 실측(1.5.0->1.6.0)과 똑같은 마커 상태: target + fail_count 만 있고 사유가 없다.
            updater._save_state({"target": "9.9.9", "fail_count": 1,
                                 "last_attempt": time.time()})
            reason = updater.failure_reason("9.9.9")
            app._log(f"[v1.7.0 확인] 사유가 마커에 없을 때(v1.6.0 이 '사유 미기록' 을 찍던 "
                     f"바로 그 상태) 실측으로 만든 사유: {reason[:300]}")

            # 다운로드 도중 프로그램이 닫힌 상황 = 실패가 아니라 '중단'. 백오프가 걸리면 안 된다.
            updater._save_state({})
            updater._set_phase("9.9.9", "download")
            info = updater._reconcile_interrupted_attempt()
            app._log(f"[v1.7.0 확인] 다운로드 중 프로그램 종료 -> 하드실패={info['hard']} "
                     f"(백오프 없음, 즉시 재시도={updater._should_attempt('9.9.9')}) "
                     f"사유='{updater._load_state()['last_error'][:120]}'")
            app._log(f"[v1.7.0 확인] 백오프 상한: "
                     f"{[updater._backoff_for(n) // 60 for n in range(1, 6)]}분 "
                     f"(하드 실패 {updater.MAX_HARD_FAILURES}회면 자동 교체 포기하고 "
                     f"매 실행 화면에 알림)")
        finally:
            updater._STATE_PATH, updater._TRAIL_PATH, updater.remote_log = keep
            import shutil as _sh
            _sh.rmtree(demo_dir, ignore_errors=True)

        # 막힌 상태의 배너를 실제로 띄워 스크린샷에 남긴다(고객이 보게 될 바로 그 화면).
        app._on_update_blocked(
            "9.9.9",
            "데모: 자동 교체가 막혔을 때 고객이 보는 화면 (실제 사유가 여기에 그대로 뜬다)",
            config.MANUAL_DOWNLOAD_URL)
    except Exception as e:
        app._log(f"[v1.7.0 확인] 실행 중 오류: {e}")

    hold_ms = int(os.environ.get("DIAG_HOLD_MS", "6000"))
    root.lift()
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass
    root.after(hold_ms, root.destroy)


def main():
    if os.environ.get("DIAG_AUTO") == "1":
        root = tk.Tk()
        app = App(root)
        root.after(300, root.destroy)
        root.mainloop()
        return 0

    root = tk.Tk()
    app = App(root)

    if "--guidemo" in sys.argv:
        root.after(500, lambda: _run_guidemo(root, app))

    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
