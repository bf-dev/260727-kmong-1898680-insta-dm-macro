# -*- coding: utf-8 -*-
"""excel_reader 단위테스트: 비연속 컬럼(C/F만) + 헤더 자동 스킵 + F열 없는 행 스킵."""

import os
import tempfile
import unittest

import openpyxl

import excel_reader


def _build_sheet(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # 헤더 행 (실제 고객 시트처럼 A/B/D/E/G 등 다른 컬럼도 존재)
    ws["A1"] = "번호"; ws["B1"] = "이름"; ws["C1"] = "인스타그램URL"
    ws["D1"] = "메모"; ws["E1"] = "담당자"; ws["F1"] = "DM문구"; ws["G1"] = "섭외메시지 캡처"

    # row2: 정상 데이터
    ws["A2"] = 1; ws["B2"] = "홍길동"; ws["C2"] = "https://www.instagram.com/user_one/"
    ws["F2"] = "안녕하세요 user_one님"; ws["G2"] = "(캡처 이미지 링크 - 무시되어야 함)"

    # row3: F열(메시지) 비어있음 -> 스킵 대상
    ws["A3"] = 2; ws["C3"] = "https://www.instagram.com/user_two/"

    # row4: 완전 빈 줄
    # (아무 것도 안 씀)

    # row5: 쿼리스트링 붙은 URL, 정상 처리돼야 함
    ws["C5"] = "https://www.instagram.com/user_three/?hl=ko"
    ws["F5"] = "user_three 님, 반갑습니다!"

    wb.save(path)


class ExcelReaderTests(unittest.TestCase):
    def setUp(self):
        fd, self.path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        _build_sheet(self.path)

    def tearDown(self):
        try:
            os.unlink(self.path)
        except Exception:
            pass

    def test_only_c_and_f_used_header_skipped(self):
        rows, skipped = excel_reader.load_rows(self.path)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].row_no, 2)
        self.assertEqual(rows[0].username, "user_one")
        self.assertEqual(rows[0].message, "안녕하세요 user_one님")
        self.assertEqual(rows[1].row_no, 5)
        self.assertEqual(rows[1].username, "user_three")  # 쿼리스트링 제거됨
        self.assertIn(3, skipped)  # F열 없는 행은 스킵 목록에

    def test_ignores_other_columns(self):
        rows, _ = excel_reader.load_rows(self.path)
        # G열(섭외메시지 캡처) 내용이 message 에 절대 섞이면 안 된다
        for r in rows:
            self.assertNotIn("캡처", r.message)


if __name__ == "__main__":
    unittest.main()
