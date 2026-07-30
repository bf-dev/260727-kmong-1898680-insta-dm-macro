# -*- coding: utf-8 -*-
"""팔로우->DM 흐름 + 재시작 이어하기(resume) + 랜덤 대기 실행을 실제 selenium 없이 검증한다.

FakeDriver 로 인스타그램 프로필/DM 페이지의 최소 DOM 을 흉내내, instagram_actions.py 와
macro_engine.py 의 '진짜 코드'(테스트 전용 스텁이 아니라 배포되는 그 함수들)를 그대로 실행한다.
확인하는 것:
  1) follow_profile 이 Follow 버튼을 실제로 찾아 클릭한다.
  2) send_dm 이 /direct/t/<username>/ 로 이동해 메시지 입력창을 찾아 사람처럼(한 글자씩) 입력하고
     Enter 로 전송한다.
  3) macro_engine 이 각 행을 순서대로(C->팔로우->F->다음) 처리하고, 완료된 행은
     progress_store 에 저장되어 재시작 시 중복 DM 되지 않는다(resume).
  4) config 의 랜덤 대기 코드 경로가 예외 없이 실행된다(범위를 테스트용으로 짧게 monkeypatch).
"""

import os
import sys
import tempfile
import time
import unittest

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

_TMP = tempfile.mkdtemp(prefix="insta_dm_test_")
os.environ["APPDATA"] = _TMP

import config
config.APP_DIR = _TMP
config.PROGRESS_DIR = os.path.join(_TMP, "progress")
# 테스트 속도를 위해 대기 범위를 짧게 (코드 경로 자체는 그대로 실행됨)
config.DELAY_AFTER_FOLLOW_MIN = 0.01
config.DELAY_AFTER_FOLLOW_MAX = 0.03
config.DELAY_BETWEEN_PEOPLE_MIN = 0.02
config.DELAY_BETWEEN_PEOPLE_MAX = 0.05
config.TYPE_JITTER_MIN = 0.001
config.TYPE_JITTER_MAX = 0.002
config.WAIT_TIMEOUT = 3
config.PROFILE_VIEW_PAUSE_MIN = 0.01
config.PROFILE_VIEW_PAUSE_MAX = 0.02
config.DM_PAGE_LOAD_PAUSE_MIN = 0.01
config.DM_PAGE_LOAD_PAUSE_MAX = 0.02
config.PRE_TYPE_PAUSE_MIN = 0.01
config.PRE_TYPE_PAUSE_MAX = 0.02
config.POST_TYPE_PAUSE_MIN = 0.01
config.POST_TYPE_PAUSE_MAX = 0.02
config.POST_FOLLOW_CLICK_PAUSE_MIN = 0.01
config.POST_FOLLOW_CLICK_PAUSE_MAX = 0.02
config.POST_SEND_PAUSE_MIN = 0.01
config.POST_SEND_PAUSE_MAX = 0.02

import bridge
bridge.remote_log = lambda *a, **k: None   # 네트워크 호출 없이 오프라인 테스트
bridge.upload_run = lambda *a, **k: None

import excel_reader
import instagram_actions as ig
import macro_engine
import progress_store


class FakeElement:
    def __init__(self, tag="div", text="", attrs=None, on_click=None, on_send_keys=None,
                 displayed=True, enabled=True):
        self.tag = tag
        self.text = text
        self.attrs = attrs or {}
        self._on_click = on_click
        self._on_send_keys = on_send_keys
        self._displayed = displayed
        self._enabled = enabled

    def get_attribute(self, name):
        return self.attrs.get(name)

    def click(self):
        if self._on_click:
            self._on_click()

    def send_keys(self, *args):
        if self._on_send_keys:
            for a in args:
                self._on_send_keys(a)

    def is_displayed(self):
        return self._displayed

    def is_enabled(self):
        return self._enabled


class FakeDriver:
    """프로필 페이지(팔로우 버튼) + DM 스레드 페이지(메시지 입력창)만 흉내내는 최소 fake."""

    def __init__(self):
        self.current_url = ""
        self.followed_usernames = []
        self.sent_messages = {}  # username -> full typed text
        self._typed_buffer = []
        self._not_found_usernames = set()

    # ---- 네비게이션 ----
    def get(self, url):
        self.current_url = url

    def execute_script(self, script):
        if "readyState" in script:
            return "complete"
        return "{}"

    def get_cookies(self):
        return []

    def get_screenshot_as_png(self):
        return b"\x89PNG\r\n"

    @property
    def page_source(self):
        return f"<html><body>fake page for {self.current_url}</body></html>"

    def _page_kind(self):
        if "/direct/t/" in self.current_url:
            return "dm"
        if "instagram.com/" in self.current_url:
            return "profile"
        return "other"

    def _current_username(self):
        if self._page_kind() == "dm":
            return self.current_url.rstrip("/").split("/direct/t/")[-1]
        if self._page_kind() == "profile":
            return self.current_url.rstrip("/").split("instagram.com/")[-1].split("?")[0]
        return None

    # ---- DOM 조회 ----
    def find_element(self, by, value):
        els = self.find_elements(by, value)
        if not els:
            raise NoSuchElementException(f"not found: {by}={value}")
        return els[0]

    def find_elements(self, by, value):
        kind = self._page_kind()
        username = self._current_username()

        if by == By.TAG_NAME and value == "header":
            return [FakeElement(tag="header")] if kind == "profile" else []

        if by == By.TAG_NAME and value == "body":
            return [FakeElement(tag="body", text="")]

        if by == By.XPATH and "role='button'" in value:
            if kind != "profile":
                return []
            if username in self._not_found_usernames:
                return []
            already = username in self.followed_usernames

            def _do_follow(u=username):
                self.followed_usernames.append(u)

            if already:
                return [FakeElement(tag="button", text="Following")]
            return [FakeElement(tag="button", text="Follow", on_click=_do_follow)]

        if by == By.XPATH and "contenteditable" in value:
            if kind != "dm":
                return []
            if username in self._not_found_usernames:
                return []
            buf = []

            def _record(ch):
                buf.append(ch)
                if ch == Keys.RETURN:
                    self.sent_messages[username] = "".join(x for x in buf if x != Keys.RETURN)

            return [FakeElement(tag="textarea", attrs={"aria-label": "Message"},
                                on_send_keys=_record)]

        return []


def _make_excel(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "번호"; ws["C1"] = "인스타그램URL"; ws["F1"] = "DM문구"
    rows = [
        (2, "https://www.instagram.com/alice_test/", "안녕하세요 alice 님, 협업 제안드려요"),
        (3, "https://www.instagram.com/bob_test/", "반갑습니다 bob 님!"),
        (4, "https://www.instagram.com/carol_test/", "carol 님께 인사드립니다"),
    ]
    for row_no, url, msg in rows:
        ws.cell(row=row_no, column=3, value=url)
        ws.cell(row=row_no, column=6, value=msg)
    wb.save(path)


class FakeApiActions:
    """ig_api 와 같은 모양의 가짜 액션 모듈. 차단/실패 시나리오를 결정적으로 재현한다."""

    SELECTOR_MISS_DETAILS = frozenset()

    def __init__(self, block_on_row=None, fail_rows=(), fail_detail="dm_rejected"):
        self.block_on_row = block_on_row
        self.fail_rows = set(fail_rows)
        self.fail_detail = fail_detail
        self.followed = []
        self.sent = {}
        self.restriction = None

    def _row_no_of(self, username):
        return {"alice_test": 2, "bob_test": 3, "carol_test": 4}.get(username)

    def follow_profile(self, session, profile_url, log=print):
        username = profile_url.rstrip("/").split("/")[-1]
        if self._row_no_of(username) == self.block_on_row:
            self.restriction = "action_block:feedback_required"
            return ig.ActionResult(False, "follow_blocked")
        self.followed.append(username)
        return ig.ActionResult(True, "followed")

    def send_dm(self, session, username, message, log=print):
        if self._row_no_of(username) in self.fail_rows:
            return ig.ActionResult(False, self.fail_detail)
        self.sent[username] = message
        return ig.ActionResult(True, "sent")

    def detect_restriction(self, session):
        return self.restriction


class MacroFlowTests(unittest.TestCase):
    def setUp(self):
        fd, self.xlsx_path = tempfile.mkstemp(suffix=".xlsx", dir=_TMP)
        os.close(fd)
        _make_excel(self.xlsx_path)
        progress_store.reset(self.xlsx_path, "test_account")
        progress_store.reset_daily_count("test_account")

    def test_follow_then_dm_single_row(self):
        driver = FakeDriver()
        result = ig.follow_profile(driver, "https://www.instagram.com/alice_test/", log=lambda *_: None)
        self.assertTrue(result.ok)
        self.assertEqual(result.detail, "followed")
        self.assertIn("alice_test", driver.followed_usernames)

        dm_result = ig.send_dm(driver, "alice_test", "안녕하세요 alice 님, 협업 제안드려요",
                               log=lambda *_: None)
        self.assertTrue(dm_result.ok)
        self.assertEqual(driver.sent_messages.get("alice_test"),
                         "안녕하세요 alice 님, 협업 제안드려요")

    def test_already_following_is_skipped_not_reclicked(self):
        driver = FakeDriver()
        driver.followed_usernames.append("alice_test")  # 이미 팔로우 중이라고 가정
        result = ig.follow_profile(driver, "https://www.instagram.com/alice_test/", log=lambda *_: None)
        self.assertTrue(result.ok)
        self.assertEqual(result.detail, "already_following")

    def test_macro_engine_processes_rows_in_order_and_persists_progress(self):
        driver = FakeDriver()
        rows, skipped = excel_reader.load_rows(self.xlsx_path)
        self.assertEqual(len(rows), 3)

        logs = []
        engine = macro_engine.MacroEngine(
            driver, rows, self.xlsx_path, "test_account",
            log_cb=logs.append)
        engine.run()  # 스레드 대신 동기 실행(테스트 결정성)

        self.assertEqual(driver.followed_usernames, ["alice_test", "bob_test", "carol_test"])
        self.assertEqual(set(driver.sent_messages.keys()),
                         {"alice_test", "bob_test", "carol_test"})
        self.assertEqual(engine.stats["followed"], 3)
        self.assertEqual(engine.stats["dm_sent"], 3)
        self.assertEqual(engine.stats["failed"], 0)

        done = progress_store.load_done_rows(self.xlsx_path, "test_account")
        self.assertEqual(done, {2, 3, 4})

    def test_resume_skips_already_done_rows_no_duplicate_dm(self):
        driver = FakeDriver()
        rows, _ = excel_reader.load_rows(self.xlsx_path)

        # 1차 실행: alice 만 처리됐다고 가정하고 진행상황을 미리 심어둔다(중간에 꺼진 상황 재현).
        progress_store.mark_done(self.xlsx_path, "test_account", 2)

        engine = macro_engine.MacroEngine(driver, rows, self.xlsx_path, "test_account",
                                          log_cb=lambda *_: None)
        engine.run()

        # alice 는 이미 완료 처리돼 있었으므로 이번 실행에서 다시 팔로우/DM 되면 안 된다.
        self.assertNotIn("alice_test", driver.followed_usernames)
        self.assertNotIn("alice_test", driver.sent_messages)
        # bob/carol 은 이번에 처리됨
        self.assertIn("bob_test", driver.followed_usernames)
        self.assertIn("carol_test", driver.followed_usernames)
        done = progress_store.load_done_rows(self.xlsx_path, "test_account")
        self.assertEqual(done, {2, 3, 4})

    def test_message_box_not_found_does_not_crash_batch(self):
        driver = FakeDriver()
        driver._not_found_usernames.add("bob_test")  # bob 은 DM 입력창을 못 찾는 상황 시뮬레이션
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        engine = macro_engine.MacroEngine(driver, rows, self.xlsx_path, "test_account",
                                          log_cb=lambda *_: None)
        engine.run()
        # bob 실패해도 alice/carol 은 정상 처리되어야 한다(배치 전체가 멈추면 안 됨)
        self.assertIn("alice_test", driver.sent_messages)
        self.assertIn("carol_test", driver.sent_messages)
        self.assertNotIn("bob_test", driver.sent_messages)
        self.assertEqual(engine.stats["failed"], 1)
        # 실패도 '명시적 실패'라 재시도 대상이 아니라 완료 처리되어(중복 DM 방지) 있어야 한다
        done = progress_store.load_done_rows(self.xlsx_path, "test_account")
        self.assertEqual(done, {2, 3, 4})


class DailyCapTests(unittest.TestCase):
    """하루 상한: 상한에 닿으면 스스로 멈추고, 남은 사람은 다음 날 이어서 처리돼야 한다."""

    def setUp(self):
        fd, self.xlsx_path = tempfile.mkstemp(suffix=".xlsx", dir=_TMP)
        os.close(fd)
        _make_excel(self.xlsx_path)
        progress_store.reset(self.xlsx_path, "cap_account")
        progress_store.reset_daily_count("cap_account")

    def test_stops_at_daily_cap_and_leaves_rest_for_tomorrow(self):
        actions = FakeApiActions()
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        engine = macro_engine.MacroEngine(object(), rows, self.xlsx_path, "cap_account",
                                          log_cb=lambda *_: None, daily_cap=2, actions=actions)
        engine.run()

        self.assertEqual(engine.halt_reason, "daily_cap")
        self.assertEqual(len(actions.sent), 2)  # 3명 중 2명만
        self.assertEqual(progress_store.get_daily_count("cap_account"), 2)
        # 처리한 2명만 완료 처리되어야 한다(나머지는 내일 이어서)
        self.assertEqual(len(progress_store.load_done_rows(self.xlsx_path, "cap_account")), 2)

    def test_already_at_cap_does_nothing(self):
        progress_store.bump_daily_count("cap_account", 5)
        actions = FakeApiActions()
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        engine = macro_engine.MacroEngine(object(), rows, self.xlsx_path, "cap_account",
                                          log_cb=lambda *_: None, daily_cap=5, actions=actions)
        engine.run()
        self.assertEqual(engine.halt_reason, "daily_cap")
        self.assertEqual(actions.sent, {})
        self.assertEqual(actions.followed, [])

    def test_daily_count_is_per_account_not_per_excel(self):
        progress_store.reset_daily_count("acc_a")
        progress_store.reset_daily_count("acc_b")
        progress_store.bump_daily_count("acc_a", 3)
        self.assertEqual(progress_store.get_daily_count("acc_a"), 3)
        self.assertEqual(progress_store.get_daily_count("acc_b"), 0)


class HaltOnBlockTests(unittest.TestCase):
    """차단 감지: 감지 즉시 멈추고, 그 행은 완료 처리하지 않아야 한다(다음에 재시도)."""

    def setUp(self):
        fd, self.xlsx_path = tempfile.mkstemp(suffix=".xlsx", dir=_TMP)
        os.close(fd)
        _make_excel(self.xlsx_path)
        progress_store.reset(self.xlsx_path, "block_account")
        progress_store.reset_daily_count("block_account")

    def test_action_block_halts_batch_and_does_not_consume_the_row(self):
        actions = FakeApiActions(block_on_row=3)  # bob 처리 중 차단
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        halts = []
        engine = macro_engine.MacroEngine(object(), rows, self.xlsx_path, "block_account",
                                          log_cb=lambda *_: None, daily_cap=50,
                                          halt_cb=lambda r, m: halts.append(r), actions=actions)
        engine.run()

        self.assertEqual(engine.halt_reason, "action_block:feedback_required")
        self.assertEqual(halts, ["action_block:feedback_required"])
        # alice 는 처리됐고, bob 에서 멈췄으므로 carol 은 손도 대지 않아야 한다.
        self.assertIn("alice_test", actions.sent)
        self.assertNotIn("bob_test", actions.sent)
        self.assertNotIn("carol_test", actions.sent)
        # 차단된 bob 행은 완료 처리되면 안 된다(다음 실행에 다시 시도해야 하므로).
        done = progress_store.load_done_rows(self.xlsx_path, "block_account")
        self.assertEqual(done, {2})
        self.assertEqual(progress_store.get_daily_count("block_account"), 1)

    def test_consecutive_failures_halt(self):
        actions = FakeApiActions(fail_rows=(2, 3, 4))
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        engine = macro_engine.MacroEngine(object(), rows, self.xlsx_path, "block_account",
                                          log_cb=lambda *_: None, daily_cap=50, actions=actions)
        engine.run()
        self.assertEqual(engine.halt_reason, "consecutive_failures")
        self.assertEqual(engine.stats["failed"], config.CONSECUTIVE_FAILURE_HALT)

    def test_halt_message_covers_every_reason(self):
        for reason in ("daily_cap", "challenge", "logged_out", "account_suspended",
                       "action_block:feedback_required", "consecutive_failures",
                       "selector_drift"):
            self.assertNotIn("사유:", macro_engine.halt_message(reason), reason)


class SelectorMissTests(unittest.TestCase):
    """셀렉터 미스: 일반 실패와 구분해 selector-miss 로 따로 올리고, 연속되면 멈춘다."""

    def setUp(self):
        fd, self.xlsx_path = tempfile.mkstemp(suffix=".xlsx", dir=_TMP)
        os.close(fd)
        _make_excel(self.xlsx_path)
        progress_store.reset(self.xlsx_path, "sel_account")
        progress_store.reset_daily_count("sel_account")

    def test_selector_miss_uploads_its_own_kind(self):
        uploaded = []
        original = bridge.upload_run
        bridge.upload_run = lambda summary, zip_path=None, kind="run": uploaded.append(kind)
        try:
            driver = FakeDriver()
            driver._not_found_usernames.add("bob_test")
            rows, _ = excel_reader.load_rows(self.xlsx_path)
            engine = macro_engine.MacroEngine(driver, rows, self.xlsx_path, "sel_account",
                                              log_cb=lambda *_: None, daily_cap=50)
            engine.run()
        finally:
            bridge.upload_run = original
        self.assertIn("selector-miss", uploaded)
        self.assertIsNone(engine.halt_reason)  # 1건만으로는 멈추지 않는다

    def test_repeated_selector_miss_halts(self):
        driver = FakeDriver()
        for u in ("alice_test", "bob_test", "carol_test"):
            driver._not_found_usernames.add(u)
        rows, _ = excel_reader.load_rows(self.xlsx_path)
        engine = macro_engine.MacroEngine(driver, rows, self.xlsx_path, "sel_account",
                                          log_cb=lambda *_: None, daily_cap=50)
        engine.run()
        self.assertEqual(engine.halt_reason, "selector_drift")


class RestrictionDetectionTests(unittest.TestCase):
    """웹 UI 엔진의 차단 화면 감지 - 평범한 화면을 차단으로 오판하면 안 된다."""

    class _Page:
        def __init__(self, url, text):
            self.current_url = url
            self._text = text

        def find_element(self, by, value):
            return FakeElement(tag="body", text=self._text)

    def test_detects_action_block_text(self):
        page = self._Page("https://www.instagram.com/someone/", "작업이 차단됨\n나중에 다시 시도해 주세요")
        self.assertTrue(str(ig.detect_restriction(page)).startswith("action_block"))

    def test_detects_challenge_url(self):
        page = self._Page("https://www.instagram.com/challenge/?next=/", "")
        self.assertEqual(ig.detect_restriction(page), "challenge")

    def test_detects_logged_out(self):
        page = self._Page("https://www.instagram.com/accounts/login/?next=/", "")
        self.assertEqual(ig.detect_restriction(page), "logged_out")

    def test_normal_profile_page_is_not_a_block(self):
        page = self._Page("https://www.instagram.com/alice_test/",
                          "게시물 120 팔로워 3,400 팔로우 180 안녕하세요 프로필 소개입니다 팔로우 메시지")
        self.assertIsNone(ig.detect_restriction(page))

    def test_long_page_is_not_scanned_for_false_positives(self):
        page = self._Page("https://www.instagram.com/alice_test/", "댓글 " * 3000 + "try again later")
        self.assertIsNone(ig.detect_restriction(page))


if __name__ == "__main__":
    unittest.main()
